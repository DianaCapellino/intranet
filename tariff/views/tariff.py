from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, FileResponse
from django.urls import reverse
from tariff.models import Location, SupplierGroup, Supplier, ProductGroup, Product, FixedRateCost, RateGroup, Rate, CostItem, RateLine, CsvFileTourplan, CsvFormTourplan, TourplanLine, Change, TYPE_HISTORY
from tariff.utils import apply_client_margin
from intranet.utils import report_tariff_error_hotel, send_templated_email, report_tariff_error_service
from intranet.models import Client, Holidays
import csv
from django.contrib.auth.decorators import login_required
from datetime import date, datetime, timedelta
import logging
from django.conf import settings
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.http import JsonResponse
from django.db.models import Q, Avg
from django.core.paginator import Paginator
import json
from collections import defaultdict
from django.contrib import messages
import math
import re

@login_required
def index(request):

    this_year = date.today().year

    return render(request, "tariff/tariff.html", {
        "locations": Location.objects.all().order_by("order"),
        "clients": Client.objects.all(),
        "this_year": this_year,
    })

logger = logging.getLogger(__name__)


def get_filtered_rate_lines(request):


    loc_id = request.GET.get('location')
    t_type = request.GET.get('type')
    season = request.GET.get('season')
    client_id = request.GET.get('client')

    if t_type == "acc":
        rate_lines = (
            RateLine.objects
            .select_related(
                "group__product__supplier__group",
                "group__product__group",
                "group__product__supplier",
            )
            .prefetch_related("line_rates")
            .order_by(
                "group__product__supplier__group__order",
                "group__product__supplier__id",
                "group__product__group__order",
                "date_from",
                "date_to",
                "group__product__order",
            )
        )
    else:
        rate_lines = (
            RateLine.objects
            .select_related(
                "group__product__supplier__group",
                "group__product__group",
                "group__product__supplier",
            )
            .prefetch_related("line_rates")
            .order_by(
                "group__product__supplier__group__order",
                "group__product__supplier__order",
                "group__product__group__order",
                "group__product__order",
                "date_from",
                "date_to",
            )
        )

    if loc_id:
        rate_lines = rate_lines.filter(group__product__group__location_id=loc_id)

    if t_type == "acc":
        rate_lines = rate_lines.filter(group__product__type_service="AC")
    else:
        rate_lines = rate_lines.filter(group__product__type_service="NA")

    if season:
        year = int(season)
        season_start = date(year, 5, 1)        # 01 May año
        season_end   = date(year + 1, 4, 30)   # 30 Apr año siguiente

        # Una tarifa pertenece a la temporada si se superpone con su rango,
        # es decir: empieza antes de que termine la temporada
        #           Y termina después de que empieza la temporada.
        rate_lines = rate_lines.filter(
            date_from__lte=season_end,    # empieza antes del 30 Apr
            date_to__gte=season_start,    # termina después del 01 May
        )

    if client_id:
        client = Client.objects.get(id=client_id)
        rate_lines = rate_lines.filter(group__product__clients__id=client_id)
    else:
        client = getattr(request.user, 'client', None)
        if client is None:
            try:
                client = Client.objects.get(name=request.user.other_name)
            except Client.DoesNotExist:
                client = None
        if client:
            rate_lines = rate_lines.filter(group__product__clients__id=client.id)
        else:
            rate_lines = rate_lines.none()

    is_client = request.user.userType == "Cliente"

    for line in rate_lines:
        rates = {}
        provisional_columns = set()

        for r in line.line_rates.all():
            if is_client and r.status != "Confirmed":
                continue

            sell_adjusted = apply_client_margin(
                rate=r,
                client_category=client.category,
                service_type="AC" if t_type == "acc" else "NA"
            )

            rates[r.column_options] = sell_adjusted
            if r.status == "Provisional":
                provisional_columns.add(r.column_options)

        line.rates_by_column = rates
        line.provisional_columns = provisional_columns

    return rate_lines, t_type, is_client


@login_required
def tariff_search(request):

    has_params = any(request.GET.get(key) for key in ['client', 'location', 'type', 'season'])
    if not has_params:
        return render(request, "tariff/tariff_table_partial.html", {'rate_lines': None})

    rate_lines, t_type, is_client = get_filtered_rate_lines(request)

    return render(request, "tariff/tariff_table_partial.html", {
        "rate_lines": rate_lines,
        "tariff_type": t_type,
        "is_client": is_client,
    })


def export_services_excel(request):

    has_params = any(request.GET.get(key) for key in ['client', 'location', 'type', 'season'])
    if not has_params:
        return HttpResponse("No data to export")

    rate_lines, t_type, is_client = get_filtered_rate_lines(request)
    # 👆 MISMA lógica que tu vista principal

    header_fill = PatternFill(start_color="D88775", end_color="D88775", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for line in rate_lines:
        print(line.rates_by_column)
        break

    if t_type == "acc":
        type_label = "Accommodation"
    else:
        type_label = "Services"

    loc_id = request.GET.get("location")
    season = request.GET.get("season")

    location_label = ""
    if loc_id:
        location = Location.objects.filter(id=loc_id).first()
        if location:
            location_label = location.name

    season_label = ""
    if season:
        year = int(season)
        season_label = f"{year}/{year+1}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Aliwen Tariff"

    filename = f"Aliwen Tariff - {type_label}"

    if location_label:
        filename += f" - {location_label}"

    if season_label:
        filename += f" - {season_label}"

    filename += ".xlsx"

    ws.freeze_panes = "A2"

    if t_type == "svs":
        ws.append([
            "Product",
            "Type",
            "Group",
            "From",
            "To",
            "1 Pax",
            "2 Pax",
            "3 Pax",
            "4 Pax",
            "5 Pax",
            "6 Pax",
        ])
        for line in rate_lines:
            ws.append([
                line.group.product.name,
                line.group.name,
                line.group.product.group.name,
                line.date_from,
                line.date_to,
                line.rates_by_column.get('1'),
                line.rates_by_column.get('2'),
                line.rates_by_column.get('3'),
                line.rates_by_column.get('4'),
                line.rates_by_column.get('5'),
                line.rates_by_column.get('6'),
            ])
    else:
        ws.append([
            "Supplier",
            "Room",
            "Type",
            "Condition",
            "From",
            "To",
            "SGL",
            "DBL",
        ])
        for line in rate_lines:
            ws.append([
                getattr(line.group.product.supplier, "name", ""),
                line.group.product.name,
                line.group.name,
                line.group.product.group.name,
                line.date_from,
                line.date_to,
                line.rates_by_column.get('SGL'),
                line.rates_by_column.get('DBL'),
            ])

    # Format of the header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Correct the width of the columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # Aligment of the cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Borders and more formats
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    alt_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def toggle_costs(request):
    request.session["show_costs"] = not request.session.get("show_costs", False)
    return redirect(request.META.get("HTTP_REFERER", "/"))


@login_required
def tp_mod_list(request):

    if request.method == "POST":
        form = CsvFormTourplan(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            csv_obj = CsvFileTourplan.objects.get(read=False)
            changes, stats = upload_data(csv_obj)

            if stats["rows_read"] == 0:
                messages.warning(
                    request,
                    "No se leyó ninguna tarifa del archivo. "
                    "Verificá que el formato del CSV sea correcto y que los códigos de proveedor y producto existan en el sistema."
                )
            else:
                parts = [f"Se procesaron <strong>{stats['rows_read']}</strong> tarifas."]
                if stats["up_to_date"]:
                    parts.append(f"<strong>{stats['up_to_date']}</strong> ya coinciden ✓")
                if stats["to_update"]:
                    parts.append(f"<strong>{stats['to_update']}</strong> para actualizar")
                if stats["to_add"]:
                    parts.append(f"<strong>{stats['to_add']}</strong> para agregar")
                if stats["to_delete"]:
                    parts.append(f"<strong>{stats['to_delete']}</strong> líneas sin coincidencia")

                if changes:
                    request.session["pending_changes"] = changes
                    messages.info(request, " · ".join(parts))
                else:
                    messages.success(request, " · ".join(parts) + " — Todo está al día.")

            return HttpResponseRedirect(reverse("tp_mod_list"))

    else:
        form = CsvFormTourplan()

    pending_changes = request.session.get("pending_changes", None)

    suppliers_ac = list(
        Supplier.objects.filter(group__type_service='AC')
        .order_by('name').values('id', 'name', 'update_tp')
    )
    suppliers_na = list(
        Supplier.objects.filter(group__type_service='NA')
        .order_by('name').values('id', 'name', 'update_tp')
    )

    return render(request, "tariff/tp_mod_list.html", {
        "form": form,
        "pending_changes_json": json.dumps(pending_changes) if pending_changes else None,
        "has_pending": bool(pending_changes),
        "suppliers_json": json.dumps({"AC": suppliers_ac, "NA": suppliers_na}),
    })


def apply_changes(request):
    """
    Receives { confirmed: [...], ignored_ids: [...] }
    Applies confirmed items and removes both confirmed + ignored from session.
    """
    if request.method == "POST":
        try:
            body = json.loads(request.body)
            confirmed    = body.get("confirmed", [])
            ignored_ids  = set(body.get("ignored_ids", []))

            if confirmed:
                apply_confirmed_changes(confirmed)
                n = len(confirmed)
                messages.success(request, f"Se guardaron <strong>{n}</strong> cambio{'s' if n != 1 else ''} correctamente.")

            # Remove processed items (confirmed + ignored) from session
            confirmed_ids = {item["_id"] for item in confirmed}
            removed_ids   = confirmed_ids | ignored_ids

            pending = request.session.get("pending_changes", [])
            remaining = [item for item in pending if item.get("_id") not in removed_ids]

            if remaining:
                request.session["pending_changes"] = remaining
                request.session.modified = True
            else:
                request.session.pop("pending_changes", None)

        except (json.JSONDecodeError, KeyError) as e:
            messages.error(request, f"Error al aplicar los cambios: {e}")

    return HttpResponseRedirect(reverse("tp_mod_list"))


def discard_changes(request):
    """Descarta todos los cambios pendientes."""
    request.session.pop("pending_changes", None)
    messages.warning(request, "Se descartaron todos los cambios pendientes.")
    return HttpResponseRedirect(reverse("tp_mod_list"))


@login_required
def toggle_supplier_update_tp(request):
    """Toggle update_tp for one or more suppliers."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    data = json.loads(request.body)
    supplier_ids = data.get('supplier_ids', [])
    update_tp    = bool(data.get('update_tp'))
    Supplier.objects.filter(pk__in=supplier_ids).update(update_tp=update_tp)
    return JsonResponse({'ok': True, 'updated': len(supplier_ids)})


@login_required
def tp_mod_list_services(request):
    """Upload services tariff file (Tourplan format for NA products)."""
    if request.method == 'POST':
        form = CsvFormTourplan(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            csv_obj = CsvFileTourplan.objects.filter(read=False).last()
            changes, stats = upload_data_services(csv_obj)

            if stats["rows_read"] == 0:
                messages.warning(
                    request,
                    "No se leyó ninguna tarifa del archivo. "
                    "Verificá que el formato del CSV sea correcto y que los códigos de proveedor y producto existan en el sistema."
                )
            else:
                parts = [f"Se procesaron <strong>{stats['rows_read']}</strong> tarifas de servicios."]
                if stats["up_to_date"]:
                    parts.append(f"<strong>{stats['up_to_date']}</strong> ya coinciden ✓")
                if stats["to_update"]:
                    parts.append(f"<strong>{stats['to_update']}</strong> para actualizar")
                if stats["to_add"]:
                    parts.append(f"<strong>{stats['to_add']}</strong> para agregar")
                if stats["to_delete"]:
                    parts.append(f"<strong>{stats['to_delete']}</strong> líneas sin coincidencia")

                if changes:
                    request.session["pending_changes"] = changes
                    messages.info(request, " · ".join(parts))
                else:
                    messages.success(request, " · ".join(parts) + " — Todo está al día.")
        else:
            messages.error(request, "Error al procesar el archivo.")

    return HttpResponseRedirect(reverse("tp_mod_list"))


def upload_data(csv_obj):
    today = date.today()
    result = []
    stats = {
        "rows_read":    0,
        "rows_skipped": 0,
        "up_to_date":   0,
        "to_update":    0,
        "to_add":       0,
        "to_delete":    0,
    }

    supplier_codes = set(
        Supplier.objects.filter(update_tp=True).values_list("code", flat=True)
    )
    product_map = {
        (p.supplier.code, p.code): p
        for p in Product.objects.select_related("supplier", "group").filter(supplier__update_tp=True)
    }

    csv_product_dates = set()
    csv_products_seen = set()

    ctx_supplier     = None
    ctx_supplier_obj = None
    ctx_product      = None
    ctx_price_code   = None
    ctx_date_from    = None
    ctx_date_to      = None
    ctx_date_from_db = None
    ctx_date_to_db   = None
    ctx_last_sgl_cost = 0.0
    ctx_last_dbl_cost = 0.0
    ctx_dbl_seen      = False  # True once "Double" row processed for this product/date

    # result_index[(product_pk, date_from_db, date_to_db, column_options, rate_group_id)]
    # → index in result list, so supplement rows replace the base row instead of appending.
    result_index = {}

    with open(csv_obj.file_name.path, "r") as f:
        reader = csv.reader(f, delimiter=";")

        for i, row in enumerate(reader):
            if i < 7:
                continue

            cols = row + [""] * max(0, 17 - len(row))

            c1  = cols[0].strip()
            c4  = cols[3].strip()
            c7  = cols[6].strip()
            c8  = cols[7].strip()
            c9  = cols[8].strip()
            c10 = cols[9].strip()
            c11 = cols[10].strip()
            c13 = cols[12].strip()
            c16 = cols[15].strip()
            c17 = cols[16].strip()

            # ---- inherit: supplier
            if c1:
                if c1 not in supplier_codes:
                    ctx_supplier = ctx_supplier_obj = None
                    ctx_product = ctx_price_code = None
                    ctx_date_from = ctx_date_to = None
                    continue
                ctx_supplier     = c1
                ctx_supplier_obj = Supplier.objects.get(code=c1)

            # ---- inherit: product — reset accumulators
            if c4:
                key               = (ctx_supplier, c4)
                ctx_product       = product_map.get(key)
                ctx_last_sgl_cost = 0.0
                ctx_last_dbl_cost = 0.0
                ctx_dbl_seen      = False

            # ---- inherit: price_code
            if c7:
                ctx_price_code = c7

            # ---- inherit: dates — reset accumulators on new date block
            if c8:
                ctx_last_sgl_cost = 0.0
                ctx_last_dbl_cost = 0.0
                ctx_dbl_seen      = False
                try:
                    date_obj         = datetime.strptime(c8, "%d/%m/%Y")
                    ctx_date_from    = c8
                    ctx_date_from_db = date_obj.date()
                except ValueError:
                    ctx_date_from = ctx_date_from_db = None             

            if c9:
                try:
                    ctx_date_to    = c9
                    ctx_date_to_db = datetime.strptime(c9, "%d/%m/%Y").date()
                except ValueError:
                    ctx_date_to = ctx_date_to_db = None

            if c10 == "K":
                status = "Confirmed"
            elif c10 == "P":
                status = "Provisional"
            else:
                status = "Confirmed"
                

            # ---- per-row: room type classification
            c11_upper = c11.upper()

            if c11 == "Single":
                column_options    = "SGL"
                is_supplement     = False
            elif "SGL" in c11_upper:
                # SGL supplement — accumulates onto Single
                column_options = "SGL"
                is_supplement  = True
            elif c11 == "Double":
                column_options = "DBL"
                is_supplement  = False
                ctx_dbl_seen   = True
            elif c11 == "Twin":
                # Use Twin only if no Double seen yet for this product/date
                if ctx_dbl_seen:
                    continue
                column_options = "DBL"
                is_supplement  = False
            elif "DBL" in c11_upper:
                # DBL supplement — accumulates onto Double
                column_options = "DBL"
                is_supplement  = True
            else:
                continue

            # ---- per-row: fcu
            if c13 == "Room":
                fcu = "Group"
            elif c13 == "Person":
                fcu = "Person"
            else:
                continue

            # ---- validate context
            if not ctx_product or not ctx_supplier_obj:
                continue
            if not ctx_date_from_db or not ctx_date_to_db:
                continue
            if ctx_date_from_db <= today:
                continue

            # ---- price_code validation
            PRICE_CODE_MAP = {
                "AU": ["Audley Exclusive Rates"],
                "NR": ["Net Rates - per night", "Rates per night for 1 night"],
                "DM": ["Net Rates - per night", "Rates per night for 1 night"],
                "CX": None,
                "TD": None,
            }
            if ctx_price_code not in PRICE_CODE_MAP:
                continue

            is_zero_cost       = PRICE_CODE_MAP[ctx_price_code] is None
            preferred_pg_names = PRICE_CODE_MAP[ctx_price_code]

            # ---- per-row: cost
            if is_zero_cost:
                cost = 0.0
            else:
                raw_cost = c16.replace("\xa0", "").replace(" ", "")
                if not raw_cost or set(raw_cost) <= {"-"}:
                    continue
                try:
                    cost = round(float(raw_cost.replace(",", ".")), 2)
                except ValueError:
                    continue

            # ---- accumulate SGL / DBL costs symmetrically
            if not is_zero_cost:
                if column_options == "SGL":
                    if is_supplement:
                        ctx_last_sgl_cost += cost
                    else:
                        ctx_last_sgl_cost = cost   # new base
                    cost = ctx_last_sgl_cost
                elif column_options == "DBL":
                    if is_supplement:
                        ctx_last_dbl_cost += cost
                    else:
                        ctx_last_dbl_cost = cost   # new base
                    cost = ctx_last_dbl_cost

            # ---- calculate sell_tourplan with final accumulated cost
            if is_zero_cost:
                sell_tourplan = 0
            else:
                margin = ctx_supplier_obj.margin
                if not margin or margin == 0:
                    continue
                sell_tourplan = math.ceil(cost / margin)

            # The exact amount in Supplier
            margin_num = ctx_supplier_obj.margin
            
            # The range option of the margin: High/Low/Regular
            margin_info = ctx_supplier_obj.margin_info

            # ---- track product+date combo for Delete detection
            csv_products_seen.add(ctx_product.pk)
            csv_product_dates.add((ctx_product.pk, ctx_date_from_db, ctx_date_to_db))

            # ---- resolve target RateGroup(s)
            all_rate_groups = list(RateGroup.objects.filter(product=ctx_product).order_by("order"))

            if not all_rate_groups:
                # Brand-new product — no RateGroup exists yet.
                # Emit a single Add entry; apply_confirmed_changes will create the group.
                dedup_key = (ctx_product.pk, ctx_date_from_db, ctx_date_to_db, column_options, None)
                if dedup_key not in result_index:
                    entry = {
                        "action":                "Add",
                        "current_sell_tourplan": None,
                        "current_sell":          None,
                        "current_cost":          None,
                        "rate_id":               None,
                        "rate_group_id":         None,
                    }
                    entry.update({
                        "product_code":    ctx_product.code,
                        "product_name":    str(ctx_product),
                        "product_id":      ctx_product.pk,
                        "rate_group_name": "Standard",
                        "price_code":      ctx_price_code,
                        "date_from":       ctx_date_from,
                        "date_to":         ctx_date_to,
                        "column_options":  column_options,
                        "fcu":             fcu,
                        "cost":            cost,
                        "status":          status,
                        "sell_tourplan":   sell_tourplan,
                        "sell":            sell_tourplan,
                        "margin":          margin_num,
                        "margin_info":     margin_info,
                        "season":          "To be defined",
                    })
                    entry["_id"] = stats["rows_read"]
                    stats["rows_read"] += 1
                    stats["to_add"]    += 1
                    result_index[dedup_key] = len(result)
                    result.append(entry)
                continue

            if preferred_pg_names is not None:
                if ctx_product.group.name in preferred_pg_names:
                    target_rate_groups = all_rate_groups
                else:
                    target_rate_groups = [all_rate_groups[0]]
            else:
                target_rate_groups = all_rate_groups

            # ---- classify Update / Add — one entry per RateGroup
            # Supplement rows REPLACE the base row in result (same dedup key)
            for rate_group in target_rate_groups:
                dedup_key = (
                    ctx_product.pk,
                    ctx_date_from_db,
                    ctx_date_to_db,
                    column_options,
                    rate_group.pk,
                )

                matching_rate_line = (
                    RateLine.objects.filter(
                        group=rate_group,
                        date_from=ctx_date_from_db,
                        date_to=ctx_date_to_db,
                    )
                    .select_related("group")
                    .first()
                )

                if matching_rate_line:
                    existing_rate = Rate.objects.filter(
                        rate_line=matching_rate_line,
                        column_options=column_options,
                    ).first()

                    if existing_rate:
                        if existing_rate.sell_tourplan != sell_tourplan:
                            entry = {
                                "action":                "Update",
                                "current_sell_tourplan": existing_rate.sell_tourplan,
                                "current_sell":          existing_rate.sell,
                                "current_cost":          existing_rate.cost,
                                "rate_id":               existing_rate.pk,
                                "rate_group_id":         None,
                            }
                        else:
                            # Up to date — ensure no stale entry from base row remains
                            if dedup_key in result_index:
                                idx = result_index.pop(dedup_key)
                                result.pop(idx)
                                # Re-index everything after the removed entry
                                result_index = {
                                    k: (v - 1 if v > idx else v)
                                    for k, v in result_index.items()
                                }
                                stats["to_update"] -= 1
                            stats["up_to_date"] += 1
                            continue
                    else:
                        entry = {
                            "action":                "Add",
                            "current_sell_tourplan": None,
                            "current_sell":          None,
                            "current_cost":          None,
                            "rate_id":               None,
                            "rate_group_id":         matching_rate_line.group.pk,
                        }
                else:
                    entry = {
                        "action":                "Add",
                        "current_sell_tourplan": None,
                        "current_sell":          None,
                        "current_cost":          None,
                        "rate_id":               None,
                        "rate_group_id":         rate_group.pk,
                    }

                # Common fields
                entry.update({
                    "product_code":   ctx_product.code,
                    "product_name":   str(ctx_product),
                    "product_id":     ctx_product.pk,
                    "price_code":     ctx_price_code,
                    "date_from":      ctx_date_from,
                    "date_to":        ctx_date_to,
                    "column_options": column_options,
                    "fcu":            fcu,
                    "cost":           cost,
                    "status":         status,
                    "sell_tourplan":  sell_tourplan,
                    "sell":           sell_tourplan,
                    "margin":         margin_num,
                    "margin_info":    margin_info,
                    "rate_group_name": rate_group.name,
                    "season":         "To be defined",
                })

                if dedup_key in result_index:
                    # Supplement: replace base row in-place with updated cost
                    idx = result_index[dedup_key]
                    entry["_id"] = result[idx]["_id"]   # keep original _id
                    result[idx]  = entry
                    # Adjust stats: undo the base row count, re-add for supplement
                    if entry["action"] == "Update":
                        pass  # already counted
                    # (stats already incremented when base was added — no double count)
                else:
                    # Base row (or only row): append fresh
                    entry["_id"] = stats["rows_read"]
                    stats["rows_read"] += 1

                    if entry["action"] == "Update":
                        stats["to_update"] += 1
                    else:
                        stats["to_add"] += 1

                    result_index[dedup_key] = len(result)
                    result.append(entry)

    # ---- Delete detection
    if csv_products_seen:
        orphan_lines = (
            RateLine.objects.filter(
                group__product__in=csv_products_seen,
                date_from__gt=today,
            )
            .select_related("group__product__supplier")
            .prefetch_related("line_rates")
        )

        next_id = stats["rows_read"]
        for rl in orphan_lines:
            product = rl.group.product
            combo   = (product.pk, rl.date_from, rl.date_to)
            if combo in csv_product_dates:
                continue

            rates = list(rl.line_rates.all())
            rates_summary = [
                {
                    "column_options": r.column_options,
                    "sell_tourplan":  r.sell_tourplan,
                    "sell":           r.sell,
                    "cost":           r.cost,
                }
                for r in rates
            ]

            stats["to_delete"] += 1
            result.append({
                "_id":            next_id,
                "action":         "Delete",
                "product_code":   product.code,
                "product_name":   str(product),
                "date_from":      rl.date_from.strftime("%d/%m/%Y"),
                "date_to":        rl.date_to.strftime("%d/%m/%Y"),
                "season":         rl.season,
                "rate_line_id":   rl.pk,
                "rates_summary":  rates_summary,
                "price_code": None, "column_options": None, "fcu": None,
                "cost": None, "sell_tourplan": None, "sell": None, "margin": None,
                "current_sell_tourplan": None, "current_sell": None, "current_cost": None,
                "rate_id": None, "rate_group_id": None,
            })
            next_id += 1

    csv_obj.read = True
    csv_obj.save()
    csv_obj.delete()

    return result, stats


_PXB_RE = re.compile(r'(\d+)\.PXB\s*\((\d+)-(\d+)\)', re.IGNORECASE)


def _cost_per_pax_sv(value, fcu, tax, increase, usd, exchange, pax):
    """Server-side mirror of calcCostPerPax in supplier_rates.html."""
    v = float(value or 0)
    if fcu == 'Group' and pax > 0:
        v /= pax
    v *= (1 + float(tax or 0) / 100)
    v *= (1 + float(increase or 0) / 100)
    if not usd and float(exchange or 1) > 0:
        v /= float(exchange or 1)
    return round(v * 100) / 100


def upload_data_services(csv_obj):
    """Parse a services (NA) Tourplan CSV and return (changes, stats).

    Same Rate-level comparison as upload_data (AC), with an extra branch
    for rates that have has_items=True: instead of comparing sell_tourplan,
    we find the matching CostItem (code == product.code), update its value,
    recalculate Rate.cost from all CostItems, and only update Rate.cost
    (sell prices are left untouched for services).
    """
    today = date.today()
    result = []
    stats = {
        "rows_read":    0,
        "rows_skipped": 0,
        "up_to_date":   0,
        "to_update":    0,
        "to_add":       0,
        "to_delete":    0,
    }

    supplier_codes = set(
        Supplier.objects.filter(update_tp=True, group__type_service='NA')
        .values_list("code", flat=True)
    )
    product_map = {
        (p.supplier.code, p.code): p
        for p in Product.objects.select_related("supplier", "group")
        .filter(supplier__update_tp=True, type_service='NA')
    }

    csv_product_dates = set()
    csv_products_seen = set()

    ctx_supplier     = None
    ctx_supplier_obj = None
    ctx_product      = None
    ctx_fcu          = None
    ctx_date_from    = None
    ctx_date_to      = None
    ctx_date_from_db = None
    ctx_date_to_db   = None
    ctx_service_desc = None  # service name from col 16 rows without PXB band

    result_index = {}

    with open(csv_obj.file_name.path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter=";")

        for i, row in enumerate(reader):
            if i < 5:
                continue

            cols = row + [""] * max(0, 21 - len(row))

            c3  = cols[2].strip()   # supplier code
            c5  = cols[4].strip()   # product code
            c8  = cols[7].strip()   # FCU: Person / Group
            c13 = cols[12].strip()  # date_from
            c14 = cols[13].strip()  # date_to
            c16 = cols[15].strip()  # PXB band
            c20 = cols[19].strip()  # cost value

            # ---- inherit: supplier
            if c3:
                if c3 not in supplier_codes:
                    ctx_supplier = ctx_supplier_obj = None
                    ctx_product = ctx_fcu = None
                    ctx_date_from = ctx_date_to = None
                    continue
                ctx_supplier     = c3
                ctx_supplier_obj = Supplier.objects.get(code=c3)

            # ---- inherit: product
            if c5:
                ctx_product = product_map.get((ctx_supplier, c5))

            # ---- inherit: FCU
            if c8 in ("Person", "Group"):
                ctx_fcu = c8

            # ---- inherit: dates
            if c13:
                try:
                    ctx_date_from_db = datetime.strptime(c13, "%d/%m/%Y").date()
                    ctx_date_from    = c13
                except ValueError:
                    ctx_date_from = ctx_date_from_db = None

            if c14:
                try:
                    ctx_date_to_db = datetime.strptime(c14, "%d/%m/%Y").date()
                    ctx_date_to    = c14
                except ValueError:
                    ctx_date_to = ctx_date_to_db = None

            # ---- validate context
            if not ctx_product or not ctx_supplier_obj:
                continue
            if not ctx_date_from_db or not ctx_date_to_db:
                continue
            if ctx_date_from_db <= today:
                continue
            if not ctx_fcu:
                continue

            # ---- col 16: either a CostItem service name or a PXB band
            if not c16:
                continue
            m = _PXB_RE.search(c16)

            if not m:
                # ── Service name row → update matching CostItems across all bases ──
                ctx_service_desc = c16
                raw_svc = c20.replace("\xa0", "").replace(" ", "")
                if not raw_svc or set(raw_svc) <= {"-"}:
                    continue
                try:
                    svc_cost = round(float(raw_svc.replace(",", ".")), 2)
                except ValueError:
                    continue

                svc_margin = ctx_supplier_obj.margin
                svc_margin_info = ctx_supplier_obj.margin_info
                if not svc_margin or svc_margin == 0:
                    continue

                csv_products_seen.add(ctx_product.pk)
                csv_product_dates.add((ctx_product.pk, ctx_date_from_db, ctx_date_to_db))

                for rate_group in RateGroup.objects.filter(product=ctx_product).order_by("order"):
                    rate_line = RateLine.objects.filter(
                        group=rate_group,
                        date_from=ctx_date_from_db,
                        date_to=ctx_date_to_db,
                    ).first()
                    if not rate_line:
                        continue

                    for rate in Rate.objects.prefetch_related("cost_items").filter(
                        rate_line=rate_line, has_items=True
                    ):
                        try:
                            base = int(rate.column_options)
                        except (ValueError, TypeError):
                            continue

                        matching_ci = next(
                            (ci for ci in rate.cost_items.all()
                             if ci.name.strip() == c16.strip()),
                            None,
                        )
                        if not matching_ci:
                            continue

                        # Store raw svc_cost in CostItem (no base division here).
                        # sell_tourplan divides by base for Group fcu.
                        new_ci_value = svc_cost

                        dedup_key = (ctx_product.pk, ctx_date_from_db, ctx_date_to_db,
                                     rate.column_options, rate_group.pk)

                        if round(new_ci_value, 2) == round(matching_ci.value, 2):
                            stats["up_to_date"] += 1
                            continue

                        new_rate_cost = round(sum(
                            new_ci_value if ci.pk == matching_ci.pk else ci.value
                            for ci in rate.cost_items.all()
                        ), 2)
                        # Divide by base for Group fcu when computing sell price
                        per_base_cost = (
                            round(new_rate_cost / base, 2)
                            if matching_ci.fcu == "Group" else new_rate_cost
                        )
                        new_sell_tp = math.ceil(per_base_cost / svc_margin)

                        entry = {
                            "action":                "Update",
                            "has_items":             True,
                            "ci_id":                 matching_ci.pk,
                            "ci_name":               matching_ci.name,
                            "ci_value":              new_ci_value,
                            "ci_fcu":                matching_ci.fcu,
                            "base":                  base,
                            "new_rate_cost":         new_rate_cost,
                            "current_cost":          round(matching_ci.value, 2),
                            "current_sell_tourplan": rate.sell_tourplan,
                            "current_sell":          rate.sell,
                            "rate_id":               rate.pk,
                            "rate_group_id":         None,
                            "sell_tourplan":         new_sell_tp,
                            "sell":                  rate.sell,
                            "cost":                  new_rate_cost,
                            "product_code":          ctx_product.code,
                            "product_name":          str(ctx_product),
                            "product_id":            ctx_product.pk,
                            "price_code":            None,
                            "date_from":             ctx_date_from,
                            "date_to":               ctx_date_to,
                            "column_options":        rate.column_options,
                            "fcu":                   ctx_fcu,
                            "status":                "Confirmed",
                            "margin":                svc_margin,
                            "margin_info":           svc_margin_info,
                            "rate_group_name":       rate_group.name,
                            "season":                rate_line.season,
                        }
                        if dedup_key in result_index:
                            idx = result_index[dedup_key]
                            entry["_id"] = result[idx]["_id"]
                            result[idx] = entry
                        else:
                            entry["_id"] = stats["rows_read"]
                            stats["rows_read"] += 1
                            stats["to_update"] += 1
                            result_index[dedup_key] = len(result)
                            result.append(entry)
                continue  # done with this service name row

            # ── PXB band row → update non-has_items rates (sell_tourplan) ──────
            band_min = int(m.group(2))
            band_max = int(m.group(3))
            if band_min > 6:
                continue

            # ---- parse cost (raw band value)
            raw_cost = c20.replace("\xa0", "").replace(" ", "")
            if not raw_cost or set(raw_cost) <= {"-"}:
                continue
            try:
                band_cost = round(float(raw_cost.replace(",", ".")), 2)
            except ValueError:
                continue

            margin      = ctx_supplier_obj.margin
            margin_info = ctx_supplier_obj.margin_info
            if not margin or margin == 0:
                continue

            csv_products_seen.add(ctx_product.pk)
            csv_product_dates.add((ctx_product.pk, ctx_date_from_db, ctx_date_to_db))

            all_rate_groups = list(RateGroup.objects.filter(product=ctx_product).order_by("order"))

            # ---- one entry per base covered by this band
            for base in range(band_min, min(band_max, 6) + 1):
                column_options = str(base)
                direct_cost   = round(band_cost / base, 2) if ctx_fcu == "Group" else band_cost
                sell_tourplan = math.ceil(direct_cost / margin)

                if not all_rate_groups:
                    dedup_key = (ctx_product.pk, ctx_date_from_db, ctx_date_to_db, column_options, None)
                    if dedup_key not in result_index:
                        entry = {
                            "action":                "Add",
                            "has_items":             False,
                            "current_sell_tourplan": None,
                            "current_sell":          None,
                            "current_cost":          None,
                            "rate_id":               None,
                            "rate_group_id":         None,
                            "product_code":          ctx_product.code,
                            "product_name":          str(ctx_product),
                            "product_id":            ctx_product.pk,
                            "rate_group_name":       "Standard",
                            "price_code":            None,
                            "date_from":             ctx_date_from,
                            "date_to":               ctx_date_to,
                            "column_options":        column_options,
                            "fcu":                   ctx_fcu,
                            "cost":                  direct_cost,
                            "status":                "Confirmed",
                            "sell_tourplan":         sell_tourplan,
                            "sell":                  sell_tourplan,
                            "margin":                margin,
                            "margin_info":           margin_info,
                            "season":                "To be defined",
                            "_id":                   stats["rows_read"],
                        }
                        stats["rows_read"] += 1
                        stats["to_add"]    += 1
                        result_index[dedup_key] = len(result)
                        result.append(entry)
                    continue

                for rate_group in all_rate_groups:
                    dedup_key = (ctx_product.pk, ctx_date_from_db, ctx_date_to_db, column_options, rate_group.pk)

                    matching_rate_line = (
                        RateLine.objects.filter(
                            group=rate_group,
                            date_from=ctx_date_from_db,
                            date_to=ctx_date_to_db,
                        )
                        .select_related("group")
                        .first()
                    )

                    if matching_rate_line:
                        existing_rate = (
                            Rate.objects
                            .filter(rate_line=matching_rate_line, column_options=column_options)
                            .first()
                        )

                        if existing_rate:
                            if existing_rate.has_items:
                                # Find the parent CostItem (code == product code)
                                parent_ci = next(
                                    (ci for ci in existing_rate.cost_items.all()
                                     if ci.code.strip() == ctx_product.code.strip()),
                                    None,
                                )
                                if parent_ci and round(band_cost, 2) != round(parent_ci.value, 2):
                                    ci_dedup_key = (ctx_product.pk, ctx_date_from_db, ctx_date_to_db,
                                                    column_options, rate_group.pk, parent_ci.pk)
                                    per_base = (
                                        round(band_cost / base, 2)
                                        if parent_ci.fcu == "Group" else band_cost
                                    )
                                    new_sell_tp = math.ceil(per_base / margin)
                                    ci_entry = {
                                        "action":                "Update",
                                        "has_items":             True,
                                        "ci_id":                 parent_ci.pk,
                                        "ci_name":               parent_ci.name or ctx_product.code,
                                        "ci_value":              band_cost,
                                        "ci_fcu":                parent_ci.fcu,
                                        "base":                  base,
                                        "new_rate_cost":         None,  # recalculated from DB on apply
                                        "current_cost":          round(parent_ci.value, 2),
                                        "current_sell_tourplan": existing_rate.sell_tourplan,
                                        "current_sell":          existing_rate.sell,
                                        "rate_id":               existing_rate.pk,
                                        "rate_group_id":         None,
                                        "sell_tourplan":         new_sell_tp,
                                        "sell":                  existing_rate.sell,
                                        "cost":                  band_cost,
                                        "product_code":          ctx_product.code,
                                        "product_name":          str(ctx_product),
                                        "product_id":            ctx_product.pk,
                                        "price_code":            None,
                                        "date_from":             ctx_date_from,
                                        "date_to":               ctx_date_to,
                                        "column_options":        column_options,
                                        "fcu":                   ctx_fcu,
                                        "status":                "Confirmed",
                                        "margin":                margin,
                                        "margin_info":           margin_info,
                                        "rate_group_name":       rate_group.name,
                                        "season":                matching_rate_line.season,
                                    }
                                    if ci_dedup_key in result_index:
                                        idx = result_index[ci_dedup_key]
                                        ci_entry["_id"] = result[idx]["_id"]
                                        result[idx] = ci_entry
                                    else:
                                        ci_entry["_id"] = stats["rows_read"]
                                        stats["rows_read"] += 1
                                        stats["to_update"] += 1
                                        result_index[ci_dedup_key] = len(result)
                                        result.append(ci_entry)
                                continue

                            # Normal sell_tourplan comparison
                            if existing_rate.sell_tourplan != sell_tourplan:
                                entry = {
                                    "action":                "Update",
                                    "has_items":             False,
                                    "current_sell_tourplan": existing_rate.sell_tourplan,
                                    "current_sell":          existing_rate.sell,
                                    "current_cost":          existing_rate.cost,
                                    "rate_id":               existing_rate.pk,
                                    "rate_group_id":         None,
                                    "sell_tourplan":         sell_tourplan,
                                    "sell":                  sell_tourplan,
                                    "cost":                  direct_cost,
                                }
                            else:
                                if dedup_key in result_index:
                                    idx = result_index.pop(dedup_key)
                                    result.pop(idx)
                                    result_index = {
                                        k: (v - 1 if v > idx else v)
                                        for k, v in result_index.items()
                                    }
                                    stats["to_update"] -= 1
                                stats["up_to_date"] += 1
                                continue
                        else:
                            entry = {
                                "action":                "Add",
                                "has_items":             False,
                                "current_sell_tourplan": None,
                                "current_sell":          None,
                                "current_cost":          None,
                                "rate_id":               None,
                                "rate_group_id":         matching_rate_line.group.pk,
                                "sell_tourplan":         sell_tourplan,
                                "sell":                  sell_tourplan,
                                "cost":                  direct_cost,
                            }
                    else:
                        entry = {
                            "action":                "Add",
                            "has_items":             False,
                            "current_sell_tourplan": None,
                            "current_sell":          None,
                            "current_cost":          None,
                            "rate_id":               None,
                            "rate_group_id":         rate_group.pk,
                            "sell_tourplan":         sell_tourplan,
                            "sell":                  sell_tourplan,
                            "cost":                  direct_cost,
                        }

                    entry.update({
                        "product_code":    ctx_product.code,
                        "product_name":    str(ctx_product),
                        "product_id":      ctx_product.pk,
                        "price_code":      None,
                        "date_from":       ctx_date_from,
                        "date_to":         ctx_date_to,
                        "column_options":  column_options,
                        "fcu":             ctx_fcu,
                        "status":          "Confirmed",
                        "margin":          margin,
                        "margin_info":     margin_info,
                        "rate_group_name": rate_group.name,
                        "season":          "To be defined",
                    })

                    if dedup_key in result_index:
                        idx = result_index[dedup_key]
                        entry["_id"] = result[idx]["_id"]
                        result[idx]  = entry
                    else:
                        entry["_id"] = stats["rows_read"]
                        stats["rows_read"] += 1
                        if entry["action"] == "Update":
                            stats["to_update"] += 1
                        else:
                            stats["to_add"] += 1
                        result_index[dedup_key] = len(result)
                        result.append(entry)

    # ---- Delete detection
    if csv_products_seen:
        orphan_lines = (
            RateLine.objects.filter(
                group__product__in=csv_products_seen,
                date_from__gt=today,
            )
            .select_related("group__product__supplier")
            .prefetch_related("line_rates")
        )

        next_id = stats["rows_read"]
        for rl in orphan_lines:
            product = rl.group.product
            combo   = (product.pk, rl.date_from, rl.date_to)
            if combo in csv_product_dates:
                continue

            rates_summary = [
                {
                    "column_options": r.column_options,
                    "sell_tourplan":  r.sell_tourplan,
                    "sell":           r.sell,
                    "cost":           r.cost,
                }
                for r in rl.line_rates.all()
            ]

            stats["to_delete"] += 1
            result.append({
                "_id":            next_id,
                "action":         "Delete",
                "has_items":      False,
                "product_code":   product.code,
                "product_name":   str(product),
                "date_from":      rl.date_from.strftime("%d/%m/%Y"),
                "date_to":        rl.date_to.strftime("%d/%m/%Y"),
                "season":         rl.season,
                "rate_line_id":   rl.pk,
                "rates_summary":  rates_summary,
                "price_code": None, "column_options": None, "fcu": None,
                "cost": None, "sell_tourplan": None, "sell": None, "margin": None,
                "current_sell_tourplan": None, "current_sell": None, "current_cost": None,
                "rate_id": None, "rate_group_id": None,
            })
            next_id += 1

    csv_obj.read = True
    csv_obj.save()
    csv_obj.delete()

    return result, stats


def apply_confirmed_changes(confirmed_items):

    # ── Paso 1: recolectar old_avg por rate_line ANTES de aplicar cambios ──
    rate_ids_to_update = [
        item["rate_id"]
        for item in confirmed_items
        if item.get("action") == "Update" and item.get("rate_id")
    ]
    rateline_data = {}
    if rate_ids_to_update:
        ratelines = RateLine.objects.filter(
            line_rates__id__in=rate_ids_to_update
        ).distinct()
        for rl in ratelines:
            old_avg = rl.line_rates.aggregate(avg=Avg("sell"))["avg"] or 0
            rateline_data[rl.id] = {
                "instance": rl,
                "old_avg":  old_avg,
            }

    # ── Paso 1b: para items Add, registrar ANTES si el producto ya tiene
    #            RateLines (para distinguir "New" vs "Add" después) ──────
    add_items = [
        item for item in confirmed_items if item.get("action") == "Add"
    ]
    # product_id → True si ya tenía al menos una RateLine antes de este batch
    product_had_ratelines = {}
    for item in add_items:
        rate_group_id = item.get("rate_group_id")
        if not rate_group_id:
            continue
        try:
            rg = RateGroup.objects.select_related("product").get(pk=rate_group_id)
        except RateGroup.DoesNotExist:
            continue
        product_id = rg.product_id
        if product_id not in product_had_ratelines:
            product_had_ratelines[product_id] = (
                RateLine.objects
                .filter(group__product_id=product_id)
                .exists()
            )

    # ── Paso 2: aplicar los cambios ──────────────────────────────────────
    added_rate_lines = []   # (RateLine instance, product_id)

    for item in confirmed_items:
        action        = item.get("action")
        cost          = item.get("cost", 0)
        sell_tourplan = item.get("sell_tourplan")
        sell          = item.get("sell", sell_tourplan)
        status      = item.get("status") or "Confirmed"
        margin_info = item.get("margin_info") or ""

        if action == "Update":
            if item.get("has_items"):
                # Update the CI value, then recalculate Rate.cost and
                # sell_tourplan from the real DB sum of all CostItems.
                ci_id    = item.get("ci_id")
                ci_value = item.get("ci_value")
                if ci_id and ci_value is not None:
                    CostItem.objects.filter(pk=ci_id).update(value=ci_value)
                rate_obj = (
                    Rate.objects.prefetch_related("cost_items")
                    .get(pk=item["rate_id"])
                )
                recalc_cost = round(
                    sum(ci.value for ci in rate_obj.cost_items.all()), 2
                )
                item_margin = item.get("margin") or 1
                item_base   = item.get("base") or 1
                item_ci_fcu = item.get("ci_fcu") or ""
                per_base = (
                    recalc_cost / item_base
                    if item_ci_fcu == "Group" and item_base > 0
                    else recalc_cost
                )
                recalc_sell_tp = math.ceil(per_base / item_margin)
                Rate.objects.filter(pk=item["rate_id"]).update(
                    cost=recalc_cost,
                    sell_tourplan=recalc_sell_tp,
                )
            else:
                update_fields = {
                    "cost":          cost,
                    "sell_tourplan": sell_tourplan,
                    "sell":          sell,
                    "status":        status,
                }
                if margin_info:
                    update_fields["margin"] = margin_info
                Rate.objects.filter(pk=item["rate_id"]).update(**update_fields)

        elif action == "Add":
            rate_group_id = item.get("rate_group_id")
            if not rate_group_id:
                # Brand-new product: create the RateGroup first
                product_id = item.get("product_id")
                if not product_id:
                    continue
                rg_name = item.get("rate_group_name") or "Standard"
                rate_group_obj, _ = RateGroup.objects.get_or_create(
                    product_id=product_id,
                    name=rg_name,
                    defaults={"order": 1},
                )
                rate_group_id = rate_group_obj.pk
            date_from = datetime.strptime(item["date_from"], "%d/%m/%Y").date()
            date_to   = datetime.strptime(item["date_to"],   "%d/%m/%Y").date()
            rate_line, _ = RateLine.objects.get_or_create(
                group_id=rate_group_id,
                date_from=date_from,
                date_to=date_to,
                defaults={"season": item.get("season") or "To be defined"},
            )
            # Resolve margin_info from DB if not present in session item (old data)
            if not margin_info:
                try:
                    margin_info = rate_line.group.product.supplier.margin_info
                except Exception:
                    margin_info = ""
            Rate.objects.get_or_create(
                rate_line=rate_line,
                column_options=item["column_options"],
                defaults={
                    "cost":          cost,
                    "sell_tourplan": sell_tourplan,
                    "sell":          sell,
                    "status":        status,
                    "margin":        margin_info,
                    "has_rate":      True,
                },
            )
            # Deduplicate by rate_line id
            if not any(rl.id == rate_line.id for rl, _ in added_rate_lines):
                product_id = rate_line.group.product_id
                added_rate_lines.append((rate_line, product_id))

        elif action == "Delete":
            rate_line_id = item.get("rate_line_id")
            if rate_line_id:
                RateLine.objects.filter(pk=rate_line_id).delete()

    # ── Paso 3a: Change para Update ──────────────────────────────────────
    for rl_id, data_rl in rateline_data.items():
        rl      = data_rl["instance"]
        old_avg = data_rl["old_avg"]
        new_avg = rl.line_rates.aggregate(avg=Avg("sell"))["avg"] or 0

        if old_avg == new_avg:
            continue

        if old_avg == 0 and new_avg > 0:
            previous = (
                RateLine.objects
                .filter(group=rl.group, season=rl.season)
                .exclude(id=rl.id)
                .order_by("-date_from")
                .first()
            )
            previous_avg = (
                previous.line_rates.aggregate(avg=Avg("sell"))["avg"] or 0
                if previous else 0
            )
            change_type = "Add"
            percent = (
                (new_avg - previous_avg) / previous_avg * 100
                if previous_avg != 0 else 0
            )
        else:
            change_type = "Update"
            percent = (new_avg - old_avg) / old_avg * 100

        Change.objects.create(
            rate_line=rl,
            type=change_type,
            amount=round(percent, 2),
        )

    # ── Paso 3b: Change para Add / New ───────────────────────────────────
    for rl, product_id in added_rate_lines:
        new_avg = rl.line_rates.aggregate(avg=Avg("sell"))["avg"] or 0
        if new_avg == 0:
            continue

        # Usar el snapshot previo al batch para decidir New vs Add
        had_ratelines = product_had_ratelines.get(product_id, False)

        if not had_ratelines:
            change_type = "New"
            percent     = 0
        else:
            previous = (
                RateLine.objects
                .filter(group=rl.group, season=rl.season)
                .exclude(id=rl.id)
                .order_by("-date_from")
                .first()
            )
            previous_avg = (
                previous.line_rates.aggregate(avg=Avg("sell"))["avg"] or 0
                if previous else 0
            )
            change_type = "Add"
            percent = (
                (new_avg - previous_avg) / previous_avg * 100
                if previous_avg != 0 else 0
            )

        Change.objects.create(
            rate_line=rl,
            type=change_type,
            amount=round(percent, 2),
        )


@login_required
def special_dates(request):
    return render(
        request,
        "tariff/special_holidays.html",
        {
            "pdf_url": "documents/Argentina_National_Holidays_2026.pdf",
        }
    )

@login_required
def download_holidays_pdf(request, year):
    if year == 2026:
        file_path = os.path.join(
            settings.BASE_DIR,
            "static/documents/Argentina_National_Holidays_2026.pdf"
        )
        return FileResponse(
            open(file_path, "rb"),
            as_attachment=True,
            filename="Argentina_National_Holidays_2026.pdf"
        )
    elif year == 2027:
        file_path = os.path.join(
            settings.BASE_DIR,
            "static/documents/Argentina_National_Holidays_2027.pdf"
        )
        return FileResponse(
            open(file_path, "rb"),
            as_attachment=True,
            filename="Argentina_National_Holidays_2027.pdf"
        )
    else:
        # Implement later this exception
        pass

@login_required
def history_of_changes(request):

    return render(
        request,
        "tariff/history_of_changes.html",
        {
            "types": TYPE_HISTORY,
            "is_admin": request.user.isAdmin,
        }
    )

@login_required
def report_error_hotel(request, supplier_id):

    supplier_obj = Supplier.objects.get(pk=supplier_id)

    if request.method == "POST":

        note = request.POST["note"]

        subject, email, template, context = report_tariff_error_hotel(request.user, supplier_obj, note)

        send_templated_email(subject, email, template, context)

        this_year = date.today().year

        return render(request, "tariff/tariff.html", {
            "locations":Location.objects.all(),
            "clients": Client.objects.all(),
            "this_year": this_year,
        })

    return render(request, "tariff/report_error.html", {
        "supplier": supplier_obj,
        "type": "acc",
    })

@login_required
def report_error_service(request, product_id):

    product_obj = Product.objects.get(pk=product_id)

    if request.method == "POST":

        note = request.POST["note"]

        subject, email, template, context = report_tariff_error_service(request.user, product_obj, note)

        send_templated_email(subject, email, template, context)

        this_year = date.today().year

        return render(request, "tariff/tariff.html", {
            "locations":Location.objects.all(),
            "clients": Client.objects.all(),
            "this_year": this_year,
        })

    return render(request, "tariff/report_error.html", {
        "product": product_obj,
        "type": "svs",
    })

@login_required
def history_of_changes_data(request):
    today = date.today()
    last_year = date(today.year - 1, today.month, today.day)

    changes = (
        Change.objects
        .select_related(
            "rate_line",
            "rate_line__group",
            "rate_line__group__product",
            "rate_line__group__product__supplier",
            "rate_line__group__product__supplier__group",
            "rate_line__group__product__supplier__group__location",
        )
        .filter(date__range=(last_year, today))
    )

    if request.user.userType == "Cliente":
        client = getattr(request.user, 'client', None)
        if client is None:
            try:
                client = Client.objects.get(name=request.user.other_name)
            except Client.DoesNotExist:
                client = None
        if client:
            changes = changes.filter(rate_line__group__product__clients=client)
        else:
            changes = changes.none()

    data = []
    for change in changes:
        product = change.rate_line.group.product
        if product.type_service == "AC":
            supplier_display = str(product.supplier)
        else:
            supplier_display = f"{product.supplier.group.location} - Services"

        row = {
            "id": change.id,
            "date": str(change.date),
            "supplier": supplier_display,
            "product": f"{product.name}, {change.rate_line.group.name}",
            "validity": f"{change.rate_line.date_from}/{change.rate_line.date_to}",
            "type": change.type,
            "amount": f"{change.amount} %",
        }
        data.append(row)

    return JsonResponse({"data": data})


@login_required
def hotel_comparison(request):
    client_id      = request.GET.get('client')
    date_ins       = request.GET.getlist('date_in')
    date_outs      = request.GET.getlist('date_out')
    loc_ids        = request.GET.getlist('location')
    suppliers_raw  = request.GET.getlist('supplier')
    products_raw   = request.GET.getlist('product')      # ← nuevo
    room_types_raw = request.GET.getlist('room_type')

    is_internal = request.user.userType != "Cliente"

    all_suppliers = (
        Supplier.objects
        .filter(supplier_products__type_service="AC")
        .select_related("group__location")
        .distinct()
        .order_by("name")
    )

    # ── Todos los productos AC, con supplier y location para el filtrado JS ──
    all_products = (
        Product.objects
        .filter(type_service="AC")
        .select_related("supplier", "group__location")
        .order_by("name")
    )

    clients_qs = (
        Client.objects.filter(department="AI").order_by("name")
        if is_internal else Client.objects.none()
    )

    context = {
        "locations":         Location.objects.all().order_by("order"),
        "all_suppliers":     all_suppliers,
        "all_products":      all_products,                              # ← nuevo
        "clients":           clients_qs,
        "selected_client":   int(client_id) if client_id else None,
        "is_internal":       is_internal,
        "searched":          False,
        # First-slot pre-fill
        "first_date_in":    date_ins[0]       if date_ins       else "",
        "first_date_out":   date_outs[0]      if date_outs      else "",
        "first_location":   int(loc_ids[0])   if loc_ids and loc_ids[0]       else None,
        "first_supplier":   int(suppliers_raw[0]) if suppliers_raw and suppliers_raw[0] else None,
        "first_product":    int(products_raw[0])  if products_raw and products_raw[0]   else None,  # ← nuevo
        "first_room_type":  room_types_raw[0] if room_types_raw else "DBL",
    }

    if not date_ins or not date_ins[0]:
        return render(request, "tariff/hotel_comparison.html", context)

    context["searched"] = True

    # Client margin
    client_category = "C"
    if client_id:
        try:
            client_category = Client.objects.get(id=client_id).category
        except Client.DoesNotExist:
            pass
    elif not is_internal:
        try:
            linked = getattr(request.user, 'client', None)
            if linked:
                client_category = linked.category
            else:
                client_category = Client.objects.get(name=request.user.other_name).category
        except Client.DoesNotExist:
            pass

    # Normalize slot lists — date_ins is the source of truth for slot count.
    # Using date_ins length avoids zip misalignment when optional fields
    # (product, supplier) send fewer values than date_in/date_out.
    n_slots = len(date_ins)
    def pad(lst, default='', n=n_slots):
        return (lst + [default] * n)[:n]

    date_outs      = pad(date_outs)
    loc_ids        = pad(loc_ids)
    suppliers_raw  = pad(suppliers_raw)
    products_raw   = pad(products_raw)
    room_types_raw = pad(room_types_raw, 'DBL')

    base_qs = (
        RateLine.objects
        .filter(group__product__type_service="AC")
        .select_related("group__product__supplier__group__location", "group__product__group")
        .prefetch_related("line_rates")
    )

    rows = []
    seen_keys = set()
    slot_errors = []

    for di_str, do_str, loc_id, sup_id, prod_id, rt in zip(
        date_ins, date_outs, loc_ids, suppliers_raw, products_raw, room_types_raw
    ):
        if not di_str or not do_str:
            continue
        try:
            date_in  = datetime.strptime(di_str, "%Y-%m-%d").date()
            date_out = datetime.strptime(do_str, "%Y-%m-%d").date()
        except ValueError:
            slot_errors.append(f"Invalid date: {di_str} / {do_str}")
            continue
        if date_in >= date_out:
            slot_errors.append(f"Check-out must be after check-in ({di_str}).")
            continue

        nights = []
        d = date_in
        while d < date_out:
            nights.append(d)
            d += timedelta(days=1)
        n_nights = len(nights)

        qs = base_qs.filter(
            date_from__lte=date_out - timedelta(days=1),
            date_to__gte=date_in,
        )
        if loc_id:
            qs = qs.filter(group__product__group__location_id=loc_id)
        if sup_id:
            qs = qs.filter(group__product__supplier_id=sup_id)
        if prod_id:                                                    # ← nuevo
            qs = qs.filter(group__product_id=prod_id)

        rate_index = defaultdict(dict)
        group_meta = {}

        for line in qs:
            rg = line.group
            if rg.id not in group_meta:
                group_meta[rg.id] = {"rate_group": rg, "product": rg.product, "supplier": rg.product.supplier}
            for r in line.line_rates.all():
                if r.column_options != rt:
                    continue
                if not is_internal and r.status == "Provisional":
                    continue
                sell = apply_client_margin(rate=r, client_category=client_category, service_type="AC")
                per_person = sell / 2 if rt == "DBL" else sell
                is_prov = r.status == "Provisional"
                for night in nights:
                    if line.date_from <= night <= line.date_to and night not in rate_index[rg.id]:
                        rate_index[rg.id][night] = {"value": per_person, "provisional": is_prov}

        for rg_id, rates_by_night in rate_index.items():
            dedup = (rg_id, rt, di_str, do_str, sup_id, prod_id)
            if dedup in seen_keys:
                continue
            seen_keys.add(dedup)
            meta = group_meta[rg_id]
            has_all  = all(n in rates_by_night for n in nights)
            has_prov = any(v["provisional"] for v in rates_by_night.values())
            total_val = round(sum(rates_by_night[n]["value"] for n in nights), 2) if has_all else None
            avg_night = round(total_val / n_nights, 2) if total_val is not None else None
            rows.append({
                "supplier":        meta["supplier"],
                "product":         meta["product"],
                "rate_group":      meta["rate_group"],
                "column":          rt,
                "nights":          nights,
                "rates":           rates_by_night,
                "n_nights":        n_nights,
                "total":           total_val,
                "avg_night":       avg_night,
                "has_all_nights":  has_all,
                "has_provisional": has_prov,
                "diff":            None,
            })

    if slot_errors:
        context["error"] = " ".join(slot_errors)

    all_nights = sorted({n for row in rows for n in row["nights"]})
    rows.sort(key=lambda r: (r["avg_night"] is None, r["avg_night"] or 0))

    complete = [r for r in rows if r["avg_night"] is not None]
    min_avg = min((r["avg_night"] for r in complete), default=None)

    for row in rows:
        if row["avg_night"] is not None and min_avg is not None:
            row["diff"] = round(row["avg_night"] - min_avg, 2)
        row_nights = set(row["nights"])
        row["rate_list"] = [
            row["rates"].get(n) if n in row_nights else "NA"
            for n in all_nights
        ]

    context.update({
        "rows":       rows,
        "min_avg":    min_avg,
        "all_nights": all_nights,
    })

    return render(request, "tariff/hotel_comparison.html", context)